"""run_budget_allocator.py - Main execution script for budget allocator"""
import pandas as pd
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from budget_allocator import budget_allocator
from utils import (
    calculate_project_priority, generate_allocation_explanation,
    calculate_skill_match_score, check_mandatory_skills, parse_required_skills
)
from config import PRIORITY_WEIGHTS


def read_input_excel(input_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Read projects and resources from Excel file.
    
    Args:
        input_path: Path to input Excel file
    
    Returns:
        Tuple of (projects_df, resources_df)
    """
    xls = pd.ExcelFile(input_path)
    
    if 'Projects' not in xls.sheet_names:
        raise ValueError("Excel file must have 'Projects' sheet")
    if 'Resources' not in xls.sheet_names:
        raise ValueError("Excel file must have 'Resources' sheet")
    
    projects_df = pd.read_excel(xls, 'Projects')
    resources_df = pd.read_excel(xls, 'Resources')
    
    return projects_df, resources_df


def generate_output_excel(allocations: List[Dict], projects_df: pd.DataFrame,
                         resources_df: pd.DataFrame, output_path: str):
    """Generate output Excel with multiple sheets.
    
    Args:
        allocations: List of allocation dictionaries
        projects_df: Original projects DataFrame
        resources_df: Original resources DataFrame
        output_path: Path to output Excel file
    """
    if not allocations:
        print("Warning: No allocations generated")
        return
    
    allocs_df = pd.DataFrame(allocations)
    
    # Calculate FTE for each allocation
    # Merge with resources to get monthly cost
    allocs_with_fte = allocs_df.merge(
        resources_df[['brid', 'cost_per_year']],
        left_on='resource_id',
        right_on='brid',
        how='left'
    )
    allocs_with_fte['cost_per_month'] = allocs_with_fte['cost_per_year'] / 12.0
    allocs_with_fte['fte_allocated'] = allocs_with_fte['allocated_cost'] / allocs_with_fte['cost_per_month']
    allocs_with_fte['fte_allocated'] = allocs_with_fte['fte_allocated'].fillna(0.0)
    
    # Allocations sheet (includes both real and dummy resources)
    allocations_sheet = allocs_with_fte[[
        'resource_id', 'resource_name', 'project_id', 'project_name',
        'month', 'allocated_cost', 'fte_allocated', 'priority_score', 'skill_match_score', 'explanation'
    ]].copy()
    
    # Project Summary
    # Filter out projects without effort_estimate_man_months (same logic as in allocator)
    project_summary_data = []
    for _, proj_row in projects_df.iterrows():
        pid = int(proj_row['project_id'])
        
        # Check if effort_estimate_man_months is provided (required for allocation)
        effort_estimate = proj_row.get('effort_estimate_man_months')
        if pd.isna(effort_estimate) or effort_estimate == 0 or effort_estimate == '':
            # Skip projects without effort_estimate_man_months in summary
            continue
        
        proj_allocations = allocs_df[allocs_df['project_id'] == pid]
        proj_allocations_with_fte = allocs_with_fte[allocs_with_fte['project_id'] == pid]
        
        allocated_budget = proj_allocations['allocated_cost'].sum()
        allocated_fte = proj_allocations_with_fte['fte_allocated'].sum()
        months_active = len(set(proj_allocations['month']))
        # Average FTE = total FTE-months / number of months allocated
        annual_fte_allocated = allocated_fte / months_active if months_active > 0 and allocated_fte > 0 else 0.0
        
        # Get requested_budget (bottom-up ask) - default to alloc_budget if not provided (backward compatible)
        requested_budget = float(proj_row.get('requested_budget', 0))
        if pd.isna(requested_budget) or requested_budget == 0:
            # If not provided, default to alloc_budget (backward compatibility)
            requested_budget = float(proj_row.get('alloc_budget', 0))
        
        total_budget = float(proj_row.get('alloc_budget', 0))  # Maximum budget that can be allocated (constraint)
        budget_utilization = (allocated_budget / total_budget * 100) if total_budget > 0 else 0.0
        
        # Calculate budget gap and fulfillment percentage
        budget_gap = requested_budget - allocated_budget  # Positive gap = unfunded request
        budget_fulfillment_pct = (allocated_budget / requested_budget * 100) if requested_budget > 0 else 0.0
        
        priority_score = calculate_project_priority(proj_row, PRIORITY_WEIGHTS)
        # Count real resources only (exclude dummy resources)
        real_resources = proj_allocations[~proj_allocations['resource_id'].str.startswith('DUMMY_', na=False)]
        resources_allocated = len(set(real_resources['resource_id']))
        dummy_resources_count = len(set(proj_allocations[proj_allocations['resource_id'].str.startswith('DUMMY_', na=False)]['resource_id']))
        
        effort_estimate = float(effort_estimate) if pd.notna(effort_estimate) else None
        
        # Get funding_source and driver from project row
        funding_source = str(proj_row.get('funding_source', '')).strip()
        driver = str(proj_row.get('driver', '')).strip()
        
        project_summary_data.append({
            'project_id': pid,
            'project_name': str(proj_row.get('project_name', f'Project {pid}')),
            'funding_source': funding_source,
            'driver': driver,
            'priority_score': priority_score,
            'requested_budget': requested_budget,  # Bottom-up budget ask
            'alloc_budget': total_budget,  # Maximum budget that can be allocated (constraint)
            'allocated_cost': allocated_budget,  # What was actually allocated
            'budget_gap': budget_gap,  # requested_budget - allocated_cost (positive = unfunded)
            'budget_fulfillment_pct': budget_fulfillment_pct,  # (allocated_cost / requested_budget) * 100
            'budget_utilization_pct': budget_utilization,  # (allocated_cost / alloc_budget) * 100
            'allocated_fte': allocated_fte,  # Total FTE-months allocated (man months)
            'fte_allocated_man_months': allocated_fte,  # FTE allocated in man months (same as allocated_fte)
            'annual_fte_allocated': annual_fte_allocated,  # Average FTE (allocated_fte / months_active)
            'months_active': months_active,
            'resources_allocated': resources_allocated,
            'dummy_resources_needed': dummy_resources_count,  # Number of dummy resources (new hires) needed
            'effort_estimate_man_months': effort_estimate,
            'is_efficiency_project': total_budget == 0
        })
    
    project_summary = pd.DataFrame(project_summary_data)
    
    # Resource Summary
    resource_summary_data = []
    for _, resource_row in resources_df.iterrows():
        resource_id = str(resource_row['brid'])
        resource_allocations = allocs_df[allocs_df['resource_id'] == resource_id]
        
        total_allocated_cost = resource_allocations['allocated_cost'].sum()
        annual_cost = float(resource_row.get('cost_per_year', 0))
        monthly_cost = annual_cost / 12.0 if annual_cost > 0 else 0.0
        utilization_pct = (total_allocated_cost / (monthly_cost * 12) * 100) if monthly_cost > 0 else 0.0
        
        # Calculate FTE
        # FTE = allocated_cost / monthly_cost (for each month, then sum)
        total_fte_allocated = 0.0
        if monthly_cost > 0:
            # Sum FTE across all allocations (total FTE-months)
            total_fte_allocated = resource_allocations['allocated_cost'].sum() / monthly_cost
        
        projects_allocated = len(set(resource_allocations['project_id']))
        months_active = len(set(resource_allocations['month']))
        
        # Average FTE = total FTE-months / number of months allocated
        annual_fte_allocated = total_fte_allocated / months_active if months_active > 0 and total_fte_allocated > 0 else 0.0
        
        resource_summary_data.append({
            'resource_id': resource_id,
            'resource_name': str(resource_row.get('employee_name', resource_id)),
            'total_allocated_cost': total_allocated_cost,
            'total_fte_allocated': total_fte_allocated,
            'annual_fte_allocated': annual_fte_allocated,
            'fte_cost_per_month': monthly_cost,
            'annual_cost': annual_cost,
            'utilization_pct': utilization_pct,
            'projects_allocated': projects_allocated,
            'months_active': months_active
        })
    
    resource_summary = pd.DataFrame(resource_summary_data)
    
    # Monthly View - FTE only (pivot table)
    # allocs_with_fte already calculated above for project summary
    monthly_view_fte = allocs_with_fte.pivot_table(
        index=['resource_id', 'resource_name'],
        columns='month',
        values='fte_allocated',
        aggfunc='sum',
        fill_value=0
    ).reset_index()
    
    # Calculate total cost and total FTE for each resource (sum across all months)
    resource_totals = allocs_df.groupby(['resource_id', 'resource_name']).agg({
        'allocated_cost': 'sum'
    }).reset_index()
    resource_totals.columns = ['resource_id', 'resource_name', 'total_cost']
    
    # Calculate total FTE for each resource
    resource_fte_totals = allocs_with_fte.groupby(['resource_id', 'resource_name'])['fte_allocated'].sum().reset_index()
    resource_fte_totals.columns = ['resource_id', 'resource_name', 'total_fte']
    
    # Get project names for each resource
    resource_projects = allocs_with_fte.groupby(['resource_id', 'resource_name'])['project_name'].apply(
        lambda x: ', '.join(sorted(set(x)))
    ).reset_index()
    resource_projects.columns = ['resource_id', 'resource_name', 'projects']
    
    # Merge FTE view with total cost, total FTE, and projects
    monthly_view = monthly_view_fte.merge(
        resource_totals,
        on=['resource_id', 'resource_name'],
        how='left'
    ).merge(
        resource_fte_totals,
        on=['resource_id', 'resource_name'],
        how='left'
    ).merge(
        resource_projects,
        on=['resource_id', 'resource_name'],
        how='left'
    )
    monthly_view['total_cost'] = monthly_view['total_cost'].fillna(0.0)
    monthly_view['total_fte'] = monthly_view['total_fte'].fillna(0.0)
    monthly_view['projects'] = monthly_view['projects'].fillna('')
    
    # Add resource+allocation column (resource_name with projects, total FTE and cost)
    monthly_view['resource_allocation'] = (
        monthly_view['resource_name'].astype(str) + 
        ' [' + monthly_view['projects'].astype(str) + ']' +
        ' (FTE: ' + monthly_view['total_fte'].round(2).astype(str) + 
        ', Cost: £' + monthly_view['total_cost'].round(2).astype(str) + ')'
    )
    
    # Reorder columns: resource info, resource+allocation, then months (FTE only), then totals at the end
    base_cols = ['resource_id', 'resource_name', 'resource_allocation']
    
    # Get all month columns (sorted)
    month_cols = sorted([col for col in monthly_view.columns 
                        if col not in base_cols and col not in ['total_cost', 'total_fte', 'projects']])
    
    # Build ordered column list: resource info, resource+allocation, months, then totals
    # Exclude 'projects' column as it's redundant (info is in resource_allocation)
    ordered_cols = base_cols + month_cols + ['total_fte', 'total_cost']
    
    # Only include columns that exist
    ordered_cols = [col for col in ordered_cols if col in monthly_view.columns]
    monthly_view = monthly_view[ordered_cols]
    
    # Priority Ranking
    priority_ranking = project_summary.sort_values('priority_score', ascending=False)[[
        'project_id', 'project_name', 'priority_score', 'requested_budget', 'alloc_budget', 
        'allocated_fte', 'fte_allocated_man_months', 'annual_fte_allocated', 'allocated_cost', 'budget_gap', 
        'budget_fulfillment_pct', 'budget_utilization_pct'
    ]].copy()
    priority_ranking['rank'] = range(1, len(priority_ranking) + 1)
    
    # Skill Gaps - projects that couldn't be fully allocated
    skill_gaps_data = []
    for _, proj_row in projects_df.iterrows():
        pid = int(proj_row['project_id'])
        proj_allocations = allocs_df[allocs_df['project_id'] == pid]
        proj_allocations_with_fte = allocs_with_fte[allocs_with_fte['project_id'] == pid]
        
        allocated_budget = proj_allocations['allocated_cost'].sum()
        allocated_fte = proj_allocations_with_fte['fte_allocated'].sum()
        months_active = len(set(proj_allocations['month']))
        # Average FTE = total FTE-months / number of months allocated
        annual_fte_allocated = allocated_fte / months_active if months_active > 0 and allocated_fte > 0 else 0.0
        total_budget = float(proj_row.get('alloc_budget', 0))
        
        if total_budget > 0 and allocated_budget < total_budget * 0.95:  # Less than 95% allocated
            # Check if it's due to skill constraints
            req_skills = parse_required_skills(proj_row.get('required_skills'))
            mandatory_skills = req_skills.get('mandatory', [])
            
            if mandatory_skills:
                # Count resources that can't be allocated due to mandatory skills
                cannot_allocate_count = 0
                for _, resource_row in resources_df.iterrows():
                    can_allocate, _ = check_mandatory_skills(resource_row, req_skills)
                    if not can_allocate:
                        cannot_allocate_count += 1
                
                skill_gaps_data.append({
                    'project_id': pid,
                    'project_name': str(proj_row.get('project_name', f'Project {pid}')),
                    'total_budget': total_budget,
                    'allocated_budget': allocated_budget,
                    'allocated_fte': allocated_fte,
                    'annual_fte_allocated': annual_fte_allocated,
                    'utilization_pct': (allocated_budget / total_budget * 100) if total_budget > 0 else 0.0,
                    'mandatory_skills': ', '.join(mandatory_skills),
                    'resources_without_skills': cannot_allocate_count,
                    'reason': 'Mandatory skills constraint - insufficient resources with required skills'
                })
    
    skill_gaps = pd.DataFrame(skill_gaps_data) if skill_gaps_data else pd.DataFrame(columns=[
        'project_id', 'project_name', 'total_budget', 'allocated_budget', 'allocated_fte', 'annual_fte_allocated', 'utilization_pct',
        'mandatory_skills', 'resources_without_skills', 'reason'
    ])
    
    # Underutilized Projects Report - Enhanced with resource and skill requirements
    underutilized_data = []
    for _, proj_row in projects_df.iterrows():
        pid = int(proj_row['project_id'])
        proj_allocations = allocs_df[allocs_df['project_id'] == pid]
        proj_allocations_with_fte = allocs_with_fte[allocs_with_fte['project_id'] == pid]
        
        allocated_budget = proj_allocations['allocated_cost'].sum()
        allocated_fte = proj_allocations_with_fte['fte_allocated'].sum()
        months_active = len(set(proj_allocations['month']))
        annual_fte_allocated = allocated_fte / months_active if months_active > 0 and allocated_fte > 0 else 0.0
        total_budget = float(proj_row.get('alloc_budget', 0))
        utilization_pct = (allocated_budget / total_budget * 100) if total_budget > 0 else 0.0
        
        # Only include budgeted projects that are not fully utilized (< 99.5% to account for rounding)
        if total_budget > 0 and utilization_pct < 99.5:
            budget_gap = total_budget - allocated_budget
            req_skills = parse_required_skills(proj_row.get('required_skills'))
            
            # Calculate required skills
            mandatory_and = req_skills.get('mandatory_and', [])
            mandatory_or = req_skills.get('mandatory_or', [])
            technical_and = req_skills.get('technical_and', [])
            technical_or = req_skills.get('technical_or', [])
            functional_and = req_skills.get('functional_and', [])
            functional_or = req_skills.get('functional_or', [])
            
            # Find resources that CAN be allocated (meet mandatory skills)
            eligible_resources = []
            for _, resource_row in resources_df.iterrows():
                can_allocate, match_info = check_mandatory_skills(resource_row, req_skills)
                if can_allocate:
                    eligible_resources.append({
                        'resource_id': str(resource_row['brid']),
                        'resource_name': str(resource_row.get('employee_name', '')),
                        'technical_skills': str(resource_row.get('technical_skills', '')),
                        'functional_skills': str(resource_row.get('functional_skills', '')),
                        'cost_per_year': float(resource_row.get('cost_per_year', 0))
                    })
            
            # Estimate FTE needed to fill the gap
            # Use average resource cost if we have eligible resources
            avg_resource_cost = sum(r['cost_per_year'] for r in eligible_resources) / len(eligible_resources) if eligible_resources else 0
            estimated_fte_needed = (budget_gap / avg_resource_cost) if avg_resource_cost > 0 else 0
            
            # Build skill requirements summary
            skill_requirements = []
            if mandatory_and:
                skill_requirements.append(f"Mandatory (ALL): {', '.join(mandatory_and)}")
            if mandatory_or:
                skill_requirements.append(f"Mandatory (ANY): {', '.join(mandatory_or)}")
            if technical_and:
                skill_requirements.append(f"Technical (ALL): {', '.join(technical_and)}")
            if technical_or:
                skill_requirements.append(f"Technical (ANY): {', '.join(technical_or)}")
            if functional_and:
                skill_requirements.append(f"Functional (ALL): {', '.join(functional_and)}")
            if functional_or:
                skill_requirements.append(f"Functional (ANY): {', '.join(functional_or)}")
            
            underutilized_data.append({
                'project_id': pid,
                'project_name': str(proj_row.get('project_name', f'Project {pid}')),
                'total_budget': total_budget,
                'allocated_budget': allocated_budget,
                'budget_gap': budget_gap,
                'utilization_pct': utilization_pct,
                'allocated_fte': allocated_fte,
                'annual_fte_allocated': annual_fte_allocated,
                'estimated_fte_needed': estimated_fte_needed,
                'eligible_resources_count': len(eligible_resources),
                'total_resources_count': len(resources_df),
                'required_skills': '; '.join(skill_requirements) if skill_requirements else 'None specified',
                'months_active': months_active,
                'resources_allocated': len(set(proj_allocations['resource_id']))
            })
    
    underutilized_projects = pd.DataFrame(underutilized_data) if underutilized_data else pd.DataFrame(columns=[
        'project_id', 'project_name', 'total_budget', 'allocated_budget', 'budget_gap', 'utilization_pct',
        'allocated_fte', 'annual_fte_allocated', 'estimated_fte_needed', 'eligible_resources_count',
        'total_resources_count', 'required_skills', 'months_active', 'resources_allocated'
    ])
    
    # Efficiency Projects
    efficiency_projects = project_summary[project_summary['is_efficiency_project'] == True][[
        'project_id', 'project_name', 'allocated_cost', 'allocated_fte', 'annual_fte_allocated', 'resources_allocated', 'months_active'
    ]].copy()
    
    # Write to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        allocations_sheet.to_excel(writer, sheet_name='Allocations', index=False)
        project_summary.to_excel(writer, sheet_name='Project_Summary', index=False)
        resource_summary.to_excel(writer, sheet_name='Resource_Summary', index=False)
        monthly_view.to_excel(writer, sheet_name='Monthly_View', index=False)
        priority_ranking.to_excel(writer, sheet_name='Priority_Ranking', index=False)
        skill_gaps.to_excel(writer, sheet_name='Skill_Gaps', index=False)
        efficiency_projects.to_excel(writer, sheet_name='Efficiency_Projects', index=False)
        underutilized_projects.to_excel(writer, sheet_name='Underutilized_Projects', index=False)
    
    # Generate Gantt chart visualization and embed in Excel
    generate_gantt_chart(allocations_sheet, resources_df, output_path, projects_df)
    
    print(f"\nOutput generated: {output_path}")
    print(f"  - {len(allocations_sheet)} allocations")
    print(f"  - {len(project_summary)} projects")
    print(f"  - {len(resource_summary)} resources")
    print(f"  - {len(efficiency_projects)} efficiency projects")
    print(f"  - {len(skill_gaps)} projects with skill gaps")
    print(f"  - {len(underutilized_projects)} underutilized projects")
    print(f"  - Gantt chart embedded in Excel")


def generate_gantt_chart(allocations_df: pd.DataFrame, resources_df: pd.DataFrame, excel_path: str, projects_df: pd.DataFrame):
    """Generate Excel native Gantt-style chart showing project and resource allocations.
    
    Args:
        allocations_df: DataFrame with allocation data
        resources_df: DataFrame with resource data
        excel_path: Path to Excel file to add the chart to
        projects_df: DataFrame with project data (for pivot table)
    """
    if allocations_df.empty:
        print("Warning: No allocations to visualize")
        return
    
    # Merge with resources to get monthly costs for percentage calculation
    allocations_with_fte = allocations_df.merge(
        resources_df[['brid', 'cost_per_year']],
        left_on='resource_id',
        right_on='brid',
        how='left'
    )
    allocations_with_fte['cost_per_month'] = allocations_with_fte['cost_per_year'] / 12.0
    allocations_with_fte['fte_allocated'] = allocations_with_fte['allocated_cost'] / allocations_with_fte['cost_per_month']
    allocations_with_fte['fte_allocated'] = allocations_with_fte['fte_allocated'].fillna(0.0)
    
    # Merge with projects to get funding_source and driver
    allocations_with_fte = allocations_with_fte.merge(
        projects_df[['project_id', 'funding_source', 'driver']],
        left_on='project_id',
        right_on='project_id',
        how='left'
    )
    allocations_with_fte['funding_source'] = allocations_with_fte['funding_source'].fillna('')
    allocations_with_fte['driver'] = allocations_with_fte['driver'].fillna('')
    
    # Calculate percentage allocation for each resource-month
    allocations_with_fte['pct_allocation'] = 0.0
    for idx, row in allocations_with_fte.iterrows():
        resource_id = row['resource_id']
        resource_row = resources_df[resources_df['brid'] == resource_id]
        if not resource_row.empty:
            resource_monthly_cost = resource_row.iloc[0]['cost_per_year'] / 12.0
            if resource_monthly_cost > 0:
                allocations_with_fte.at[idx, 'pct_allocation'] = (row['allocated_cost'] / resource_monthly_cost * 100)
    
    # Get unique months
    months = sorted(allocations_with_fte['month'].unique())
    
    # Create grouped data structure: funding_source -> driver -> project -> resources
    grouped_data = []
    
    # Group by funding_source first
    funding_sources = sorted(allocations_with_fte['funding_source'].unique())
    
    for funding_source in funding_sources:
        fs_allocations = allocations_with_fte[allocations_with_fte['funding_source'] == funding_source]
        
        # Group by driver within funding_source
        drivers = sorted(fs_allocations['driver'].unique())
        
        for driver in drivers:
            driver_allocations = fs_allocations[fs_allocations['driver'] == driver]
            
            # Group by project within driver
            projects = sorted(driver_allocations['project_name'].unique())
            
            for project_name in projects:
                proj_allocations = driver_allocations[driver_allocations['project_name'] == project_name]
                resources_in_project = sorted(proj_allocations['resource_name'].unique())
                
                project_resources = []
                for resource_name in resources_in_project:
                    resource_allocations = proj_allocations[proj_allocations['resource_name'] == resource_name]
                    
                    row_data = []
                    for month in months:
                        month_allocations = resource_allocations[resource_allocations['month'] == month]
                        if not month_allocations.empty:
                            # Sum percentage allocations for this resource-month
                            pct = month_allocations['pct_allocation'].sum()
                            row_data.append(min(pct, 100.0))  # Cap at 100%
                        else:
                            row_data.append(0.0)
                    
                    project_resources.append({
                        'resource_name': resource_name,
                        'data': row_data
                    })
                
                grouped_data.append({
                    'funding_source': funding_source,
                    'driver': driver,
                    'project_name': project_name,
                    'resources': project_resources
                })
    
    # Load the Excel workbook
    try:
        wb = load_workbook(excel_path)
        
        # Create or clear the Gantt chart sheet
        if 'Gantt_Chart' in wb.sheetnames:
            ws = wb['Gantt_Chart']
            wb.remove(ws)
        
        ws = wb.create_sheet('Gantt_Chart')
        
        # Add title row
        title_cell = ws['A1']
        title_cell.value = 'Resource Allocation Heatmap (Grouped by Funding Source > Driver > Project)\n(Percentage of resource capacity allocated per month)'
        ws.merge_cells(f'A1:{chr(65 + len(months))}1')
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # Write header row (row 2)
        ws.cell(row=2, column=1, value='Funding Source / Driver / Project / Resource')
        for col_idx, month in enumerate(months, start=2):
            cell = ws.cell(row=2, column=col_idx, value=month)
            # Format header cells
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Format the label column header
        label_header = ws.cell(row=2, column=1)
        label_header.fill = header_fill
        label_header.font = header_font
        label_header.alignment = Alignment(horizontal='center')
        
        ws.row_dimensions[2].height = 25
        
        # Write grouped data: funding_source > driver > project > resources
        current_row = 3
        data_start_row = 3
        data_end_row = 2
        
        # Color scheme: funding_source (darkest), driver (medium), project (lightest)
        funding_source_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')  # Dark blue
        funding_source_font = Font(bold=True, size=12, color='FFFFFF')
        
        driver_fill = PatternFill(start_color='8FAADC', end_color='8FAADC', fill_type='solid')  # Medium blue
        driver_font = Font(bold=True, size=11, color='000000')
        
        project_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # Light blue
        project_font = Font(bold=True, size=10, color='000000')
        
        current_funding_source = None
        current_driver = None
        
        for group in grouped_data:
            funding_source = group['funding_source']
            driver = group['driver']
            project_name = group['project_name']
            resources = group['resources']
            
            # Add funding_source header if changed
            if funding_source != current_funding_source:
                if current_funding_source is not None:
                    current_row += 1  # Gap between funding sources
                
                funding_source_header = ws.cell(row=current_row, column=1, value=f'Funding Source: {funding_source if funding_source else "N/A"}')
                funding_source_header.fill = funding_source_fill
                funding_source_header.font = funding_source_font
                funding_source_header.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(f'A{current_row}:{chr(65 + len(months))}{current_row}')
                ws.row_dimensions[current_row].height = 25
                current_row += 1
                current_funding_source = funding_source
                current_driver = None  # Reset driver when funding source changes
            
            # Add driver header if changed
            if driver != current_driver:
                driver_header = ws.cell(row=current_row, column=1, value=f'  Driver: {driver if driver else "N/A"}')
                driver_header.fill = driver_fill
                driver_header.font = driver_font
                driver_header.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(f'A{current_row}:{chr(65 + len(months))}{current_row}')
                ws.row_dimensions[current_row].height = 22
                current_row += 1
                current_driver = driver
            
            # Add project header row
            project_header = ws.cell(row=current_row, column=1, value=f'    Project: {project_name}')
            project_header.fill = project_fill
            project_header.font = project_font
            project_header.alignment = Alignment(horizontal='left', vertical='center')
            
            # Merge project header across all columns
            ws.merge_cells(f'A{current_row}:{chr(65 + len(months))}{current_row}')
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
            # Add resource rows for this project
            for resource in resources:
                resource_name = resource['resource_name']
                row_data = resource['data']
                
                ws.cell(row=current_row, column=1, value=f'      {resource_name}')  # Indent resource names
                for col_idx, value in enumerate(row_data, start=2):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    # Format as percentage with 0 decimal places
                    cell.number_format = '0"%"'
                
                current_row += 1
            
            # Add a small gap row between projects (optional, for better visual separation)
            # Don't add gap after last project
            if group != grouped_data[-1]:
                # Check if next group is different funding_source or driver
                next_group = grouped_data[grouped_data.index(group) + 1]
                if next_group['funding_source'] != funding_source or next_group['driver'] != driver:
                    current_row += 1  # Gap between different funding sources or drivers
            data_end_row = current_row - 1
        
        # Apply conditional formatting (heatmap) to data cells (exclude project header rows)
        # We need to apply formatting only to resource rows, not project header rows
        # Color scale: Green (0%) -> Yellow (50%) -> Red (100%)
        data_range = f"B{data_start_row}:{chr(65 + len(months))}{data_end_row}"
        
        # Create color scale rule: Green (low) -> Yellow (mid) -> Red (high)
        color_scale = ColorScaleRule(
            start_type='num',
            start_value=0,
            start_color='90EE90',  # Light green
            mid_type='num',
            mid_value=50,
            mid_color='FFFF00',  # Yellow
            end_type='num',
            end_value=100,
            end_color='FF6B6B'  # Light red
        )
        
        ws.conditional_formatting.add(data_range, color_scale)
        
        # Add color scale legend
        legend_row = data_end_row + 3
        ws.cell(row=legend_row, column=1, value='Color Scale:')
        ws.cell(row=legend_row, column=2, value='0%').fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        ws.cell(row=legend_row, column=3, value='50%').fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws.cell(row=legend_row, column=4, value='100%').fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        ws.cell(row=legend_row, column=5, value='(Green = Low, Yellow = Medium, Red = High)')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        for col_idx in range(2, len(months) + 2):
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = 12
        
        # Freeze panes for easier navigation (freeze title and header rows)
        ws.freeze_panes = 'B3'
        
        # Add borders to cells for better readability
        from openpyxl.styles import Border, Side
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add borders to header row
        for col in range(1, len(months) + 2):
            ws.cell(row=2, column=col).border = thin_border
        
        # Add borders to all data rows (including project headers and resource rows)
        for row in range(3, data_end_row + 1):
            for col in range(1, len(months) + 2):
                cell = ws.cell(row=row, column=col)
                # Check if it's a header row (funding_source, driver, or project)
                is_funding_source_header = (cell.fill.start_color.rgb == funding_source_fill.start_color.rgb if cell.fill and cell.fill.start_color else False)
                is_driver_header = (cell.fill.start_color.rgb == driver_fill.start_color.rgb if cell.fill and cell.fill.start_color else False)
                is_project_header = (cell.fill.start_color.rgb == project_fill.start_color.rgb if cell.fill and cell.fill.start_color else False)
                
                if is_funding_source_header:
                    # Thickest border for funding source
                    thick_bottom = Border(left=Side(style='thin'), right=Side(style='thin'),
                                          top=Side(style='thin'), bottom=Side(style='thick'))
                    cell.border = thick_bottom
                elif is_driver_header:
                    # Medium border for driver
                    medium_bottom = Border(left=Side(style='thin'), right=Side(style='thin'),
                                           top=Side(style='thin'), bottom=Side(style='medium'))
                    cell.border = medium_bottom
                elif is_project_header:
                    # Thin-medium border for project
                    thin_medium_bottom = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.border = thin_medium_bottom
                else:
                    # For resource rows, add thin borders
                    cell.border = thin_border
        
        # Create pivot table sheet with underlying data
        create_pivot_table_sheet(wb, allocations_with_fte, projects_df, months)
        
        # Save the workbook
        wb.save(excel_path)
        wb.close()
        
        print(f"Excel native Gantt chart created in: {excel_path}")
    except Exception as e:
        print(f"Warning: Could not create Excel chart: {e}")
        import traceback
        traceback.print_exc()


def create_pivot_table_sheet(wb, allocations_with_fte: pd.DataFrame, projects_df: pd.DataFrame, months: List[str]):
    """Create a pivot table sheet with the underlying data used for the heatmap.
    
    Args:
        wb: Workbook object
        allocations_with_fte: DataFrame with allocations and FTE data
        projects_df: DataFrame with project data
        months: List of month strings
    """
    # Create pivot table data
    pivot_data = []
    
    for _, alloc_row in allocations_with_fte.iterrows():
        resource_id = alloc_row['resource_id']
        resource_name = alloc_row['resource_name']
        project_id = alloc_row['project_id']
        project_name = alloc_row['project_name']
        month = alloc_row['month']
        allocated_cost = alloc_row['allocated_cost']
        fte_allocated = alloc_row['fte_allocated']
        pct_allocation = alloc_row['pct_allocation']
        
        # Get project metadata
        proj_row = projects_df[projects_df['project_id'] == project_id]
        if not proj_row.empty:
            funding_source = proj_row.iloc[0].get('funding_source', '')
            driver = proj_row.iloc[0].get('driver', '')
        else:
            funding_source = ''
            driver = ''
        
        pivot_data.append({
            'Funding Source': funding_source,
            'Driver': driver,
            'Project ID': project_id,
            'Project Name': project_name,
            'Resource ID': resource_id,
            'Resource Name': resource_name,
            'Month': month,
            'Allocated Cost': allocated_cost,
            'FTE Allocated': fte_allocated,
            'Pct Allocation': pct_allocation
        })
    
    pivot_df = pd.DataFrame(pivot_data)
    
    # Create pivot table sheet
    if 'Pivot_Table' in wb.sheetnames:
        ws_pivot = wb['Pivot_Table']
        wb.remove(ws_pivot)
    
    ws_pivot = wb.create_sheet('Pivot_Table')
    
    # Write pivot data
    for col_idx, col_name in enumerate(pivot_df.columns, start=1):
        ws_pivot.cell(row=1, column=col_idx, value=col_name)
        # Format header
        header_cell = ws_pivot.cell(row=1, column=col_idx)
        header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_cell.font = Font(bold=True, color='FFFFFF')
        header_cell.alignment = Alignment(horizontal='center')
    
    # Write data rows
    for row_idx, (_, row_data) in enumerate(pivot_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(pivot_df.columns, start=1):
            value = row_data[col_name]
            cell = ws_pivot.cell(row=row_idx, column=col_idx, value=value)
            
            # Format percentage column
            if col_name == 'Pct Allocation':
                cell.number_format = '0.00"%"'
            elif col_name in ['Allocated Cost', 'FTE Allocated']:
                if col_name == 'Allocated Cost':
                    cell.number_format = '£#,##0.00'
                else:
                    cell.number_format = '0.00'
    
    # Adjust column widths
    column_widths = {
        'Funding Source': 20,
        'Driver': 20,
        'Project ID': 12,
        'Project Name': 30,
        'Resource ID': 15,
        'Resource Name': 25,
        'Month': 12,
        'Allocated Cost': 15,
        'FTE Allocated': 12,
        'Pct Allocation': 15
    }
    
    for col_idx, col_name in enumerate(pivot_df.columns, start=1):
        if col_name in column_widths:
            ws_pivot.column_dimensions[chr(64 + col_idx)].width = column_widths[col_name]
    
    # Freeze header row
    ws_pivot.freeze_panes = 'A2'
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, len(pivot_df) + 2):
        for col in range(1, len(pivot_df.columns) + 1):
            ws_pivot.cell(row=row, column=col).border = thin_border


def main():
    """Main execution function."""
    if len(sys.argv) < 2:
        print("Usage: python run_budget_allocator.py <input_excel_file> [output_excel_file]")
        print("\nExample:")
        print("  python run_budget_allocator.py excel/sample_input.xlsx excel/output.xlsx")
        sys.exit(1)
    
    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else input_path.replace('.xlsx', '_output.xlsx')
    
    if not Path(input_path).exists():
        print(f"Error: Input file not found: {input_path}")
        sys.exit(1)
    
    print(f"Reading input from: {input_path}")
    
    try:
        # Read input
        projects_df, resources_df = read_input_excel(input_path)
        print(f"  - {len(projects_df)} projects")
        print(f"  - {len(resources_df)} resources")
        
        # Run allocator
        print("\nRunning budget allocator...")
        allocations = budget_allocator(resources_df, projects_df)
        print(f"  - Generated {len(allocations)} allocations")
        
        # Generate output
        print("\nGenerating output...")
        generate_output_excel(allocations, projects_df, resources_df, output_path)
        
        print("\n✓ Allocation complete!")
        
    except Exception as e:
        print(f"\n✗ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
